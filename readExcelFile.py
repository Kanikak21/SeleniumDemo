import xlrd
import pandas as pd
from openpyxl import load_workbook


df = pd.read_excel('Login.xlsx')
print(df)
print(df.shape[0])

# workbook= xlrd.open_workbook("Login.xlsx")
# sheet= workbook.sheet_by_name("Sheet1")
# rowCount= sheet.nrows
# print(rowCount)


# workbook = load_workbook('Login.xlsx')
# worksheet = workbook.active
# for row in worksheet.iter_rows(values_only=True):
#     print(row)
# workbook.close()